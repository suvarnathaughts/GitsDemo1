import openpyxl
# this is to give location of our excel file"PythonDemo"
book = openpyxl.load_workbook("C:\\Users\\bhama\\OneDrive\\Documents\\PythonDemo .xlsx")
# This is to get control on active sheet i.e. sheet1
sheet = book.active
# This is to read from the Excel sheet
cell = sheet.cell(row=1, column=2)
print(cell.value)
# This is to write in to Excel sheet
sheet.cell(row=2, column=2).value = "Rahul"
# to see which value is writen to row2 column2
print(sheet.cell(row=2, column=2).value)
# to know how many total no.of rows present in sheet
print(sheet.max_row)
# for column
print(sheet.max_column)
# if we want value of specific cell by shortcut e.g. A5 means 5th row 1st column =t4
print(sheet['A5'].value)
print(sheet['B1'].value)
# we will see how to use for loop
# to get 1st column value
#for i in range(1, sheet.max_row+1):
    #print(sheet.cell(row=i, column=1).value)
# to print all column and row value the for loop will be
#for i in range(1, sheet.max_row+1):
    #for j in range(1, sheet.max_column + 1):
        #print(sheet.cell(row=i, column=j).value)

# if we want to print only Testcase2 value and complete data under it
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column+1):  # here we use 1 to print Testcase2,d,e,g and 2 to print only d,e,g
            print(sheet.cell(row=i, column=j).value)





        print


